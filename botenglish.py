import os
import re
import random
import asyncio
import aiohttp
import time
import math
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from aiogram import Bot, Dispatcher, Router, F
from aiogram.filters import Command, StateFilter
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    CallbackQuery,
)
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage


load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
MYMEMORY_EMAIL = os.getenv("MYMEMORY_EMAIL")
BASE_DIR = Path(__file__).resolve().parent


PAGE_SIZE = 20

# --- translate cache (очень простой) ---
TR_CACHE: dict[tuple[str, str], str] = {}   # key=(src_lang, text)

# --- simple rate limit for /tr ---
TR_LAST_TS: dict[int, float] = {}           # key=user_id -> last_ts
TR_MIN_INTERVAL = 2.0  # seconds

# ==== НАСТРОЙКИ ФАЙЛА ====
FILE_PATH = str(BASE_DIR / "vocab.xlsx")

SHEET_NAME = "THINK L2 DUTCH"   # если будет ошибка листа — поставь None
# SHEET_NAME = None

router = Router()
VOCAB: list[dict] = []
VOCAB_BY_ID: dict[int, dict] = {}


def tr_rate_limited(user_id: int) -> bool:
    now = time.time()
    last = TR_LAST_TS.get(user_id, 0.0)
    if now - last < TR_MIN_INTERVAL:
        return True
    TR_LAST_TS[user_id] = now
    return False


# ===================== Excel =====================
def load_vocab_openpyxl(path: str, sheet_name: str | None) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    headers = []
    for cell in ws[1]:
        val = cell.value if cell.value is not None else ""
        headers.append(str(val).strip())

    # В твоём файле 1-я колонка пустая — там слово
    if headers and headers[0] == "":
        headers[0] = "WORD"

    rows: list[dict] = []
    _id = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        item: dict = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            item[h] = row[i] if i < len(row) else None

        word = str(item.get("WORD") or "").strip()
        if not word or word.lower() in ("none", "nan"):
            continue

        item["ID"] = _id
        _id += 1

        unit_raw = item.get("UNIT NO")
        try:
            item["UNIT NO"] = int(unit_raw)
        except Exception:
            item["UNIT NO"] = 0

        for k in ["DEFINITION", "DUTCH TRANSLATION", "PoS", "EXAMPLE SENTENCE", "EXAMPLE"]:
            if k in item and item[k] is not None:
                item[k] = str(item[k]).strip()

        rows.append(item)

    return rows


def _clean_text(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    if s.lower() in ("none", "nan"):
        return ""
    return s


def format_items(items: list[dict]) -> str:
    out = []
    for it in items:
        _id = it.get("ID", "")
        word = _clean_text(it.get("WORD"))
        dutch = _clean_text(it.get("DUTCH TRANSLATION"))
        pos = _clean_text(it.get("PoS"))
        definition = _clean_text(it.get("DEFINITION"))
        example = _clean_text(it.get("EXAMPLE SENTENCE")) or _clean_text(it.get("EXAMPLE"))

        extra = []
        if dutch:
            extra.append(dutch)
        if pos:
            extra.append(pos)
        extra_txt = f" ({', '.join(extra)})" if extra else ""

        block = [f"{_id}. {word}{extra_txt}", f"— {definition}"]
        if example:
            block.append(f"💬 Example: {example}")

        out.append("\n".join(block))

    return "\n\n".join(out)


async def send_long(m: Message, text: str, chunk: int = 3500):
    for i in range(0, len(text), chunk):
        await m.answer(text[i:i + chunk])


# ===================== Translation (RU/EN -> HY) =====================
def detect_source_lang(text: str) -> str:
    t = (text or "").strip()
    if not t:
        return "en"

    # Armenian
    for ch in t:
        code = ord(ch)
        if 0x0530 <= code <= 0x058F:
            return "hy"

    # Cyrillic -> RU
    for ch in t:
        code = ord(ch)
        if 0x0400 <= code <= 0x04FF:
            return "ru"

    return "en"


async def translate_to_armenian(text: str) -> str:
    text = (text or "").strip()
    if not text:
        return "Напиши текст 🙂"

    src = detect_source_lang(text)
    if src == "hy":
        return text

    cache_key = (src, text)
    if cache_key in TR_CACHE:
        return TR_CACHE[cache_key]


    url = "https://api.mymemory.translated.net/get"
    params = {"q": text, "langpair": f"{src}|hy"}
    if MYMEMORY_EMAIL:
        params["de"] = MYMEMORY_EMAIL

    async with aiohttp.ClientSession() as session:
        async with session.get(url, params=params, timeout=25) as r:
            r.raise_for_status()
            data = await r.json()

    translated = ((data.get("responseData") or {}).get("translatedText")) or ""
    translated = translated.strip()
    TR_CACHE[cache_key] = translated or "Не получилось перевести 😕"
    return TR_CACHE[cache_key]


# ===================== FSM =====================
class TranslateState(StatesGroup):
    waiting_text = State()


class QuizState(StatesGroup):
    waiting_units = State()
    waiting_mode = State()
    waiting_count = State()
    in_quiz = State()


# ===================== Keyboards =====================
def build_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📚 Units"), KeyboardButton(text="🔎 Find")],
            [KeyboardButton(text="📄 Range help"), KeyboardButton(text="🧩 Unit help")],
            [KeyboardButton(text="🇦🇲 Перевод"), KeyboardButton(text="🧪 Тест")],
            [KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
    )


def build_mode_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="📝 WORD → Definition", callback_data="quizmode:wd"),
            ],
            [
                InlineKeyboardButton(text="🧩 Definition → WORD", callback_data="quizmode:dw"),
            ],
            [
                InlineKeyboardButton(text="❌ Stop", callback_data="quiz:stop"),
            ],
        ]
    )


def build_quiz_answers_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="A", callback_data="quizans:0"),
                InlineKeyboardButton(text="B", callback_data="quizans:1"),
                InlineKeyboardButton(text="C", callback_data="quizans:2"),
            ],
            [InlineKeyboardButton(text="❌ Stop test", callback_data="quiz:stop")],
        ]
    )

def build_unit_page_kb(unit_no: int, page: int, max_page: int) -> InlineKeyboardMarkup:
    buttons = []
    if page > 1:
        buttons.append(InlineKeyboardButton(text="⬅️ Назад", callback_data=f"unitpage:{unit_no}:{page-1}"))
    if page < max_page:
        buttons.append(InlineKeyboardButton(text="➡️ Дальше", callback_data=f"unitpage:{unit_no}:{page+1}"))

    rows = []
    if buttons:
        rows.append(buttons)
    rows.append([InlineKeyboardButton(text="❌ Закрыть", callback_data="unitpage:close")])

    return InlineKeyboardMarkup(inline_keyboard=rows)

async def send_unit_page(m: Message, unit_no: int, page: int):
    items = [it for it in VOCAB if it.get("UNIT NO") == unit_no]
    if not items:
        await m.answer(f"Unit {unit_no} не найден или пустой.")
        return

    total = len(items)
    max_page = max(1, math.ceil(total / PAGE_SIZE))
    page = max(1, min(page, max_page))

    start = (page - 1) * PAGE_SIZE
    end = start + PAGE_SIZE
    chunk = items[start:end]

    text = f"📘 Unit {unit_no} — страница {page}/{max_page} (слова {start+1}-{min(end, total)} из {total})\n\n"
    text += format_items(chunk)

    await send_long(m, text)
    await m.answer("Навигация:", reply_markup=build_unit_page_kb(unit_no, page, max_page))



# ===================== Bot commands =====================
@router.message(Command("start"))
async def start(m: Message):
    await m.answer(
        "Я читаю твой Excel со словами 📘 + умею переводить на армянский 🇦🇲 + тест 🧪\n\n"
        "Команды:\n"
        "/range 100 141 — слова по номерам\n"
        "/unit 4 — слова из Unit\n"
        "/find boring — поиск по слову\n"
        "/units — список unit-ов\n"
        "/tr text — перевод на армянский\n\n"
        "Кнопки: 🇦🇲 Перевод, 🧪 Тест",
        reply_markup=build_kb(),
    )


@router.message(Command("units"))
async def units_cmd(m: Message):
    counts = {}
    for it in VOCAB:
        u = it.get("UNIT NO", 0)
        if u:
            counts[u] = counts.get(u, 0) + 1

    if not counts:
        await m.answer("Units не найдены.")
        return

    text = "Units:\n" + "\n".join([f"Unit {u}: {counts[u]} words" for u in sorted(counts)])
    await m.answer(text)


@router.message(Command("range"))
async def range_cmd(m: Message):
    parts = (m.text or "").split()
    if len(parts) < 3:
        await m.answer("Пример: /range 100 141")
        return

    try:
        a = int(parts[1]); b = int(parts[2])
    except ValueError:
        await m.answer("Нужно два числа. Пример: /range 100 141")
        return

    if a > b:
        a, b = b, a

    if not VOCAB:
        await m.answer("Список слов пуст.")
        return

    a = max(a, 1)
    b = min(b, VOCAB[-1]["ID"])

    items = [it for it in VOCAB if a <= it["ID"] <= b]
    if not items:
        await m.answer("Ничего не найдено.")
        return

    await send_long(m, f"Слова {a}–{b}:\n\n" + format_items(items))

@router.message(Command("unit"))
async def unit_cmd(m: Message):
    parts = (m.text or "").split()

    if len(parts) < 2:
        await m.answer("Пример: /unit 4  или  /unit 4 2 (страница 2)")
        return

    try:
        unit_no = int(parts[1])
    except ValueError:
        await m.answer("Unit должен быть числом. Пример: /unit 4")
        return

    page = 1
    if len(parts) >= 3 and parts[2].isdigit():
        page = int(parts[2])

    await send_unit_page(m, unit_no, page)

@router.callback_query(F.data.startswith("unitpage:"))
async def unitpage_cb(cb: CallbackQuery):
    if cb.data == "unitpage:close":
        await cb.message.delete()
        await cb.answer()
        return

    _, unit_s, page_s = cb.data.split(":")
    unit_no = int(unit_s)
    page = int(page_s)

    await cb.answer()
    await cb.message.answer(f"⏭ Открываю Unit {unit_no}, страница {page}…")
    await send_unit_page(cb.message, unit_no, page)


@router.message(Command("find"))
async def find_cmd(m: Message):
    parts = (m.text or "").split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await m.answer("Пример: /find boring")
        return

    q = parts[1].strip().lower()
    items = [it for it in VOCAB if q in _clean_text(it.get("WORD")).lower()][:30]

    if not items:
        await m.answer("Не нашёл.")
        return

    await send_long(m, f"Найдено (первые {len(items)}):\n\n" + format_items(items))


# ---- buttons (подсказки) ----
@router.message(F.text == "📚 Units")
async def units_button(m: Message):
    await units_cmd(m)

@router.message(F.text == "🔎 Find")
async def find_button(m: Message):
    await m.answer("Напиши: /find boring")

@router.message(F.text == "📄 Range help")
async def range_help(m: Message):
    await m.answer("Пример: /range 100 141")

@router.message(F.text == "🧩 Unit help")
async def unit_help(m: Message):
    await m.answer("Пример: /unit 4")


# ===================== Translate (button flow) =====================
@router.message(F.text == "🇦🇲 Перевод")
async def tr_button(m: Message, state: FSMContext):
    await state.set_state(TranslateState.waiting_text)
    await m.answer("Напиши слово/текст — переведу на армянский 🇦🇲\n(или нажми ❌ Отмена)")


@router.message(Command("tr"))
async def tr_cmd(m: Message):
    if tr_rate_limited(m.from_user.id):
        await m.answer("⏳ Слишком часто. Подожди 2 секунды 🙂")
        return
    
    text = (m.text or "").split(maxsplit=1)
    if len(text) < 2 or not text[1].strip():
        await m.answer("Пример: /tr Hello world")
        return

    try:
        hy = await translate_to_armenian(text[1])
        await m.answer(f"🇦🇲 {hy}")
    except Exception as e:
        await m.answer(f"Не смог перевести 😕 ({type(e).__name__})")


@router.message(TranslateState.waiting_text)
async def tr_state_handler(m: Message, state: FSMContext):
    if tr_rate_limited(m.from_user.id):
        await m.answer("⏳ Слишком часто. Подожди 2 секунды 🙂")
        return
    if (m.text or "").startswith("/"):
        await m.answer("Если хочешь выйти — нажми ❌ Отмена. Если хочешь перевод — напиши текст без /")
        return
    try:
        hy = await translate_to_armenian(m.text or "")
        await m.answer(f"🇦🇲 {hy}")
    except Exception as e:
        await m.answer(f"Не смог перевести 😕 ({type(e).__name__})")


@router.message(StateFilter("*"), F.text.in_({"❌ Отмена", "Отмена"}))
@router.message(StateFilter("*"), Command("cancel"))
async def cancel_any(m: Message, state: FSMContext):
    await state.clear()
    await m.answer("Отменил ✅", reply_markup=build_kb())



# ===================== QUIZ =====================
def parse_quiz_source(text: str) -> tuple[str, list[int], list[int]]:
    """
    Возвращает:
      label: текст для шапки ("Units: 1,3" или "IDs: 140-160")
      pool_ids: список ID слов для теста
      units: список unit-ов (если выбраны), иначе []
    """
    t = (text or "").strip().lower()

    # --- range formats ---
    # "140-160"
    m = re.match(r"^\s*(\d+)\s*-\s*(\d+)\s*$", t)
    if not m:
        # "140 160"
        m = re.match(r"^\s*(\d+)\s+(\d+)\s*$", t)
    if not m:
        # "от 140 до 160"
        m = re.search(r"от\s*(\d+)\s*до\s*(\d+)", t)

    if m:
        a = int(m.group(1))
        b = int(m.group(2))
        if a > b:
            a, b = b, a

        pool_ids = []
        for it in VOCAB:
            _id = int(it["ID"])
            if a <= _id <= b and _clean_text(it.get("WORD")) and _clean_text(it.get("DEFINITION")):
                pool_ids.append(_id)

        label = f"IDs: {a}-{b}"
        return label, pool_ids, []

    # --- units formats (старое поведение) ---
    units = parse_units(text)
    units_set = set(units)

    pool_ids = []
    for it in VOCAB:
        if it.get("UNIT NO") in units_set:
            if _clean_text(it.get("WORD")) and _clean_text(it.get("DEFINITION")):
                pool_ids.append(int(it["ID"]))

    label = f"Units: {', '.join(map(str, units))}" if units else "Units: (none)"
    return label, pool_ids, units


def parse_units(text: str) -> list[int]:
    """
    Поддержка:
    - "1 3"
    - "1,3"
    - "1-3"
    - "1-3,5"
    """
    text = (text or "").strip()
    if not text:
        return []

    text = text.replace(",", " ").replace(";", " ")
    tokens = text.split()

    units: set[int] = set()
    for tok in tokens:
        tok = tok.strip()
        if not tok:
            continue
        if "-" in tok:
            a, b = tok.split("-", 1)
            if a.isdigit() and b.isdigit():
                a_i, b_i = int(a), int(b)
                if a_i > b_i:
                    a_i, b_i = b_i, a_i
                for u in range(a_i, b_i + 1):
                    if u > 0:
                        units.add(u)
        else:
            if tok.isdigit():
                u = int(tok)
                if u > 0:
                    units.add(u)

    return sorted(units)


def build_pool(units: list[int]) -> list[int]:
    # Берём только те слова, где есть и WORD и DEFINITION
    pool = []
    units_set = set(units)
    for it in VOCAB:
        if it.get("UNIT NO") in units_set:
            if _clean_text(it.get("WORD")) and _clean_text(it.get("DEFINITION")):
                pool.append(int(it["ID"]))
    return pool


def pick_options(correct_id: int, pool_ids: list[int], mode: str) -> tuple[list[str], int]:
    """
    mode:
      wd: options = definitions
      dw: options = words
    """
    correct_item = VOCAB_BY_ID[correct_id]
    if mode == "wd":
        correct_text = _clean_text(correct_item.get("DEFINITION"))
        get_text = lambda _id: _clean_text(VOCAB_BY_ID[_id].get("DEFINITION"))
    else:
        correct_text = _clean_text(correct_item.get("WORD"))
        get_text = lambda _id: _clean_text(VOCAB_BY_ID[_id].get("WORD"))

    # Берём 2 других ID (для вариантов) из пула, исключая correct
    others = [x for x in pool_ids if x != correct_id]
    wrong_ids = random.sample(others, 2)

    options = [correct_text, get_text(wrong_ids[0]), get_text(wrong_ids[1])]
    random.shuffle(options)
    correct_idx = options.index(correct_text)
    return options, correct_idx

async def send_next_question(m: Message, state: FSMContext):
    st = await state.get_data()

    order: list[int] = st.get("quiz_order", [])
    pos: int = st.get("quiz_pos", 0)
    score: int = st.get("quiz_score", 0)
    total: int = st.get("quiz_total", 0)
    mode: str = st.get("quiz_mode", "wd")  # "wd" или "dw"
    label: str = st.get("quiz_label", "Test")
    pool_ids: list[int] = st.get("quiz_pool_ids", [])
    wrong_ids: list[int] = st.get("quiz_wrong", [])

   # если тест закончился
    if pos >= total:
        label = st.get("quiz_label", f"Units: {', '.join(map(str, units))}" if units else "Test")
        summary = (
            f"🏁 Тест закончен!\n"
            f"{label}\n"
            f"✅ {score}/{total}"
        )

        if not wrong_ids:
            await state.clear()
            await m.answer(summary + "\n\n🔥 Ошибок нет!")
            return

        # убрать дубли, сохранив порядок
        seen = set()
        uniq_wrong = []
        for wid in wrong_ids:
            if wid not in seen:
                seen.add(wid)
                uniq_wrong.append(wid)

        lines = []
        for wid in uniq_wrong[:30]:  # покажем первые 30
            it = VOCAB_BY_ID.get(int(wid))
            if not it:
                continue
            w = _clean_text(it.get("WORD"))
            d = _clean_text(it.get("DEFINITION"))
            lines.append(f"• {w} — {d}")

        await state.clear()
        await m.answer(summary + f"\n\n❌ Ошибки ({len(uniq_wrong)}):\n" + "\n".join(lines))
        if len(uniq_wrong) > 30:
            await m.answer("…и ещё есть ошибки, но я показал первые 30.")
        return

    # следующий вопрос
    correct_id = order[pos]
    correct_item = VOCAB_BY_ID[correct_id]

    # варианты (A/B/C)
    options, correct_idx = pick_options(correct_id, pool_ids, mode)

    # сохраняем текущий вопрос и сдвигаем позицию вперёд (чтобы не повторялся)
    await state.update_data(
        quiz_correct_idx=correct_idx,
        quiz_options=options,
        quiz_current_id=correct_id,
        quiz_pos=pos + 1,
    )

    word = _clean_text(correct_item.get("WORD"))
    definition = _clean_text(correct_item.get("DEFINITION"))

    qn = pos + 1
    header = (
    f"🧪 Тест ({qn}/{total}) | Score: {score}/{pos}\n"
    f"{label}\n"
    )

    if mode == "wd":
        body = (
            f"\nСлово: {word}\n\n"
            f"A) {options[0]}\n\n"
            f"B) {options[1]}\n\n"
            f"C) {options[2]}"
        )
    else:
        body = (
            f"\nDefinition: {definition}\n\n"
            f"A) {options[0]}\n\n"
            f"B) {options[1]}\n\n"
            f"C) {options[2]}"
        )

    await m.answer(header + body, reply_markup=build_quiz_answers_kb())



@router.message(F.text == "🧪 Тест")
@router.message(Command("test"))
async def quiz_start(m: Message, state: FSMContext):
    await state.clear()
    await state.set_state(QuizState.waiting_units)
    await m.answer(
        "Напиши unit-ы для теста:\n"
        "Примеры: 6  |  1 3  |  1,3  |  1-3  |  1-3,5"
    )


@router.message(QuizState.waiting_units)
async def quiz_set_units(m: Message, state: FSMContext):
    label, pool_ids, units = parse_quiz_source(m.text or "")

    if len(pool_ids) < 3:
        await m.answer(
            "Слишком мало слов с definition для теста (нужно минимум 3).\n"
            "Примеры:\n"
            "• unit-ы: 1 3  |  1-3,5\n"
            "• диапазон: 140-160  |  140 160  |  от 140 до 160"
        )
        return

    await state.update_data(
        quiz_label=label,
        quiz_units=units,
        quiz_pool_ids=pool_ids,
    )
    await state.set_state(QuizState.waiting_mode)
    await m.answer("Выбери режим теста:", reply_markup=build_mode_kb())


@router.callback_query(F.data.startswith("quizmode:"))
async def quiz_choose_mode(cb: CallbackQuery, state: FSMContext):
    st = await state.get_data()
    if not st.get("quiz_pool_ids"):
        await cb.answer("Сначала выбери unit-ы", show_alert=True)
        return

    mode = cb.data.split(":", 1)[1]  # wd / dw
    await state.update_data(quiz_mode=mode)
    await state.set_state(QuizState.waiting_count)

    max_possible = len(st["quiz_pool_ids"])
    await cb.message.answer(f"Сколько вопросов? (1..{max_possible})\nНапример: 10")
    await cb.answer()


@router.message(QuizState.waiting_count)
async def quiz_set_count(m: Message, state: FSMContext):
    st = await state.get_data()
    pool_ids: list[int] = st["quiz_pool_ids"]
    max_possible = len(pool_ids)

    text = (m.text or "").strip().lower()
    if text.isdigit():
        total = int(text)
    else:
        total = 10  # дефолт

    total = max(1, min(total, max_possible))

    # порядок вопросов без повторов
    order = pool_ids[:]
    random.shuffle(order)
    order = order[:total]

    await state.update_data(
        quiz_order=order,
        quiz_total=total,
        quiz_pos=0,
        quiz_score=0,
        quiz_wrong=[],
    )
    await state.set_state(QuizState.in_quiz)

    await m.answer("Старт! Выбирай A / B / C")
    await send_next_question(m, state)


@router.callback_query(F.data == "quiz:stop")
async def quiz_stop(cb: CallbackQuery, state: FSMContext):
    st = await state.get_data()
    score = st.get("quiz_score", 0)
    pos = st.get("quiz_pos", 0)
    total = st.get("quiz_total", pos)
    units = st.get("quiz_units", [])
    await state.clear()
    await cb.message.answer(f"Остановил тест ✅\nUnits: {', '.join(map(str, units))}\nСчёт: ✅ {score}/{min(pos, total)}")
    await cb.answer()


@router.callback_query(F.data.startswith("quizans:"))
async def quiz_answer(cb: CallbackQuery, state: FSMContext):
    st = await state.get_data()
    if not st.get("quiz_options"):
        await cb.answer("Теста нет. Нажми 🧪 Тест.", show_alert=True)
        await state.clear()
        return

    chosen = int(cb.data.split(":")[1])
    correct_idx = st["quiz_correct_idx"]
    options = st["quiz_options"]
    mode = st.get("quiz_mode", "wd")
    current_id = st.get("quiz_current_id")

    item = VOCAB_BY_ID.get(int(current_id))
    word = _clean_text(item.get("WORD")) if item else ""
    definition = _clean_text(item.get("DEFINITION")) if item else ""

    score = st.get("quiz_score", 0)

    if chosen == correct_idx:
        score += 1
        await cb.message.answer("✅ Верно!")
    else:
        corr_letter = ["A", "B", "C"][correct_idx]
        await cb.message.answer(f"❌ Неверно. Правильный ответ: {corr_letter}")

        wrong_list = st.get("quiz_wrong", [])
        wrong_list.append(int(current_id))  # сохраняем ID слова, которое завалил
        await state.update_data(quiz_wrong=wrong_list)


    # показ правильной пары
    if mode == "wd":
        await cb.message.answer(f"{word} — {definition}")
    else:
        await cb.message.answer(f"{definition}\n— {word}")

    await state.update_data(quiz_score=score)
    await cb.answer()

    # следующий вопрос
    await send_next_question(cb.message, state)


# ===================== main =====================
async def main():
    global VOCAB, VOCAB_BY_ID

    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN not set in .env (BOT_TOKEN=...)")

    VOCAB = load_vocab_openpyxl(FILE_PATH, SHEET_NAME)
    VOCAB_BY_ID = {int(it["ID"]): it for it in VOCAB}

    bot = Bot(BOT_TOKEN)
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
# 